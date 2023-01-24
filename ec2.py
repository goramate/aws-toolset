#!/usr/bin/python3
import boto3
import pprint
from openpyxl import Workbook
MAXRESULTS=100
pp = pprint.PrettyPrinter(indent=4)
#re
def get_tag(resource,tag_name):
    if 'Tags' in resource:
        for tag in resource['Tags']:
            if tag['Key'] == tag_name:
                return  tag['Value']
    return ''

def list_ec2_instances(running_only=False):
    print('List instances')
    instances = []
    ec2client = boto3.client('ec2')
    isNextToken = True
    while(isNextToken):
        if 'NextToken'  in locals():
            response = ec2client.describe_instances(
                MaxResults=MAXRESULTS,
                NextToken= NextToken
            )
        else:
            response = ec2client.describe_instances(
                MaxResults=MAXRESULTS
            )
        for reservation in response["Reservations"]:
            for instance in reservation["Instances"]:
                if not running_only or instance['State']['Name'] == 'running':
                    instance['Name']=get_tag(instance,'Name')
                    instances.append(instance)
                #print(instance)
        if 'NextToken' in response:
            isNextToken = True
            NextToken = response['NextToken']
            #print("NextToken: " +  NextToken)
        else:
            isNextToken = False
    return instances

def list_instance_profiles():
    print('List profiles')
    profiles = {}
    iamclient = boto3.client('iam')
    isNextToken = True
    while(isNextToken):
        if 'Marker'  in locals():
            response = iamclient.list_instance_profiles(
                Marker=Marker,
                MaxItems=MAXRESULTS
            )
        else:
            response = iamclient.list_instance_profiles(
                MaxItems=MAXRESULTS
            )
        for profile in response['InstanceProfiles']:
            if not profile['Arn'] in profiles:
                profiles[profile['Arn']] = profile
        if 'Marker' in response:
            isNextToken = True
            Marker = response['Marker']
        else:
            isNextToken = False
    return profiles


def get_iam_role(IamRoleName,IamClient=''):
    if IamClient == '':
        IamClient = boto3.client('iam')
    IamRole = IamClient.get_role( RoleName=IamRoleName )
    pp.pprint(IamRole)

def get_iam_role_inline_policies(IamRoleName,IamClient=''):
    role_policies = {}
    if IamClient == '':
        IamClient = boto3.client('iam')
    IamPolicies = IamClient.list_role_policies(
        RoleName=IamRoleName,
    )
    for policy in IamPolicies['PolicyNames']:
        response = IamClient.get_role_policy(
            RoleName=IamRoleName,
            PolicyName=policy
        )
        role_policies[policy] = response['PolicyDocument']
    return role_policies

def get_iam_role_policies(IamRoleName,IamClient=''):
    role_policies = {}
    if IamClient == '':
        IamClient = boto3.client('iam')
    IamPolicies = IamClient.list_attached_role_policies(
        RoleName=IamRoleName
    )
    #pp.pprint(IamPolicies)
    for policy in IamPolicies['AttachedPolicies']:
        IamPolicy = IamClient.get_policy(
            PolicyArn = policy['PolicyArn']
        )
        #pp.pprint(IamPolicy)
        IamPolicyVersion = IamClient.get_policy_version(
            PolicyArn=policy['PolicyArn'],
            VersionId=IamPolicy['Policy']['DefaultVersionId']
        )
        #pp.pprint(IamPolicyVersion['PolicyVersion']['Document'])
        role_policies[policy['PolicyName']] = IamPolicyVersion['PolicyVersion']['Document']
    return role_policies

#generate hash instnace-id => profile
def list_instance_roles():
    instances = list_ec2_instances(running_only=True)
    profiles = list_instance_profiles()
    for profile_arn in profiles:
        profiles[profile_arn]['InstanceList'] = []
        profiles[profile_arn]['InstanceNames'] = []
    for instance in instances:
         if 'IamInstanceProfile' in instance:
            if instance['IamInstanceProfile']['Arn'] in profiles:
                 profiles[instance['IamInstanceProfile']['Arn']]['InstanceList'].append(instance['InstanceId'])
                 profiles[instance['IamInstanceProfile']['Arn']]['InstanceNames'].append(instance['Name'])
    IamClient = boto3.client('iam')
    for profile_arn in profiles:
        #pp.pprint(profile_arn)
        for role in profiles[profile_arn]['Roles']:
            role_policies = get_iam_role_inline_policies(
                IamRoleName = role['RoleName'],
                IamClient= IamClient
            )
            role['InLinePolicies'] = role_policies
        for role in profiles[profile_arn]['Roles']:
            role_policies = get_iam_role_policies(
                IamRoleName = role['RoleName'],
                IamClient= IamClient
            )
            role['Policies'] = role_policies
    pp.pprint(profiles)
    headers = ["IAMProfileName","InstanceList","InstanceNames","IamRoles","IAMPolicies","Documents"]
    values = []

    generate_report(headers,values,profiles)

def convert_array(tab):
    value = ''
    for r in tab:
        value = value + pp.pformat(r) + "\n";
    return value

def generate_report(headers,values,object):
    wb = Workbook()
    ws = wb.create_sheet("IAMProfiles")
    head_no = 0;
    col_number=len(headers)
    row_number=len(object)


    data=[]
    index=0
    for obj in object:
        #print("\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n\n")
        #pp.pprint( object[obj])
        data_row=[]
        roles = []
        policies = []
        documents = []
        for role in object[obj]['Roles']:
            roles.append(role['RoleName'])
            for policy in role['InLinePolicies']:
                policies.append(policy)
                for doc in role['InLinePolicies'][policy]['Statement']:
                    documents.append(doc)
            for policy in role['Policies']:
                policies.append(policy)
                for doc in role['Policies'][policy]['Statement']:
                    documents.append(doc)

        data_row.append(obj)
        data_row.append(convert_array(object[obj]['InstanceList']))
        #pp.pprint(object[obj]['InstanceNames'])
        data_row.append(convert_array(object[obj]['InstanceNames']))
        data_row.append(convert_array(roles))
        data_row.append(convert_array(policies))
        data_row.append(convert_array(documents))

        index+=1
        data.append(data_row)
    #header
    for col in ws.iter_cols(min_row=1, max_col=col_number, max_row=1):
        for cell in col:
            cell.value = headers[head_no]
            head_no+=1
    for row_no in range(1,row_number):
        for row in ws.iter_rows(min_row=row_no+1, max_col=col_number, max_row=row_no+1):
            head_no = 0;
            for cell in row:
                cell.value = data[row_no-1][head_no]
                head_no+=1
    wb.save("sample.xlsx")












if __name__ == '__main__':
    list_instance_roles()
